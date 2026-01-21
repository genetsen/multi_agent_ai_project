"""
Email Scanner - Core scanning and classification logic.

Orchestrates email retrieval, classification, extraction, and routing
for the Email Scanner Agent.
"""

import re
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass, field
from pathlib import Path
from enum import Enum

try:
    from .gmail_client import GmailClient, EmailMessage, EmailAttachment
except ImportError:
    from gmail_client import GmailClient, EmailMessage, EmailAttachment

logger = logging.getLogger(__name__)


class Classification(Enum):
    """Email classification result."""
    DATA_HIGH_CONFIDENCE = "DATA_EMAIL_HIGH_CONFIDENCE"
    DATA_MEDIUM_CONFIDENCE = "DATA_EMAIL_MEDIUM_CONFIDENCE"
    UNCERTAIN = "UNCERTAIN_REVIEW"
    NOT_DATA = "NOT_DATA"


@dataclass
class ClassificationResult:
    """Result of email classification."""
    classification: Classification
    score: float
    factors: List[str]
    matched_patterns: List[str] = field(default_factory=list)


@dataclass
class ExtractedAsset:
    """Represents an extracted data asset from an email."""
    asset_type: str  # "attachment", "link", "inline_table"
    source_email_id: str
    filename: Optional[str] = None
    file_path: Optional[Path] = None
    url: Optional[str] = None
    link_type: Optional[str] = None  # "google_sheets", "dropbox", etc.
    mime_type: Optional[str] = None
    size_bytes: Optional[int] = None
    row_count: Optional[int] = None
    columns: Optional[List[str]] = None
    extraction_confidence: float = 1.0
    error: Optional[str] = None


@dataclass
class PartnerPattern:
    """Configuration for a known partner pattern."""
    partner_name: str
    sender_patterns: List[str]
    subject_patterns: List[str] = field(default_factory=list)
    expected_format: Optional[str] = None
    expected_columns: Optional[List[str]] = None
    column_aliases: Dict[str, str] = field(default_factory=dict)
    auto_process: bool = True
    confidence_threshold: float = 0.8


@dataclass
class ScanResult:
    """Result of scanning a single email."""
    email: EmailMessage
    classification: ClassificationResult
    extracted_assets: List[ExtractedAsset]
    matched_partner: Optional[str] = None
    pattern_match_confidence: float = 0.0
    route_action: str = "review"  # "auto_process", "review", "skip"


class EmailClassifier:
    """
    Classifies emails based on patterns to determine if they contain data.
    """

    # High confidence subject patterns
    HIGH_CONF_SUBJECT_PATTERNS = [
        r"\b(weekly|daily|monthly|quarterly)\s+(report|data|metrics|analytics)\b",
        r"\bperformance\s+(report|data|export|metrics)\b",
        r"\b(Q[1-4]|\d{4})\s+(results|report|data)\b",
        r"\bmedia\s+(plan|report|data|metrics)\b",
        r"\b(campaign|placement|partner)\s+report\b",
    ]

    # Medium confidence subject patterns
    MEDIUM_CONF_SUBJECT_PATTERNS = [
        r"\b(attached|export|summary|numbers|results)\b",
        r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}\b",
        r"\bweek\s*(of|ending)?\s*\d",
    ]

    # Skip patterns (definitely not data)
    SKIP_SUBJECT_PATTERNS = [
        r"^(re:|fwd:|fw:)\s*re:",
        r"\bout\s+of\s+office\b",
        r"\bautomatic\s+reply\b",
        r"\bunsubscribe\b",
        r"\bcalendar\b",
        r"\bmeeting\s+(invite|invitation)\b",
        r"\b(invoice|payment|billing)\b",
    ]

    # Body patterns indicating data delivery
    DELIVERY_PHRASES = [
        "please find attached",
        "attached please find",
        "attached is the",
        "here is the data",
        "here is the report",
        "here are the numbers",
        "as requested",
        "for your review",
        "latest numbers",
        "updated report",
        "this week's data",
        "this month's report",
    ]

    # Metric keywords in body
    METRIC_PATTERNS = [
        r"\b(impressions?|imps?)\b",
        r"\b(clicks?|click-through)\b",
        r"\b(spend|cost|budget)\b",
        r"\b(conversions?|converts?)\b",
        r"\b(views?|video\s*views?)\b",
        r"\b(CPM|CPC|CPA|CPV|CPCV)\b",
        r"\b(CTR|VTR|CVR)\b",
    ]

    # Data file extensions
    DATA_EXTENSIONS = {
        ".csv": 0.30,
        ".xlsx": 0.30,
        ".xls": 0.30,
        ".tsv": 0.30,
        ".json": 0.20,
        ".xml": 0.20,
    }

    # Link patterns
    LINK_PATTERNS = {
        "google_sheets": r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)",
        "dropbox": r"(dropbox\.com/s/[a-zA-Z0-9]+|dl\.dropboxusercontent\.com)",
        "box": r"(box\.com/s/[a-zA-Z0-9]+|app\.box\.com/s/[a-zA-Z0-9]+)",
        "onedrive": r"(1drv\.ms/|onedrive\.live\.com)",
        "s3_presigned": r"s3[.-][a-z0-9-]+\.amazonaws\.com.*X-Amz-Signature",
    }

    def __init__(self):
        # Compile regex patterns for efficiency
        self._high_subj = [re.compile(p, re.I) for p in self.HIGH_CONF_SUBJECT_PATTERNS]
        self._med_subj = [re.compile(p, re.I) for p in self.MEDIUM_CONF_SUBJECT_PATTERNS]
        self._skip_subj = [re.compile(p, re.I) for p in self.SKIP_SUBJECT_PATTERNS]
        self._metric = [re.compile(p, re.I) for p in self.METRIC_PATTERNS]
        self._links = {k: re.compile(v, re.I) for k, v in self.LINK_PATTERNS.items()}

    def classify(self, email: EmailMessage) -> ClassificationResult:
        """
        Classify an email to determine if it contains data.

        Args:
            email: EmailMessage to classify

        Returns:
            ClassificationResult with score and factors
        """
        score = 0.0
        factors = []

        # Check skip patterns first
        for pattern in self._skip_subj:
            if pattern.search(email.subject):
                return ClassificationResult(
                    classification=Classification.NOT_DATA,
                    score=0.0,
                    factors=["skip_pattern_matched"],
                )

        # Subject analysis
        subj_score, subj_factors = self._analyze_subject(email.subject)
        score += subj_score
        factors.extend(subj_factors)

        # Body analysis
        body = email.body_text or email.body_html or ""
        body_score, body_factors = self._analyze_body(body)
        score += body_score
        factors.extend(body_factors)

        # Attachment analysis
        att_score, att_factors = self._analyze_attachments(email.attachments)
        score += att_score
        factors.extend(att_factors)

        # Link analysis
        link_score, link_factors = self._analyze_links(body)
        score += link_score
        factors.extend(link_factors)

        # Adjustments
        if email.subject.lower().startswith(("re:", "fwd:", "fw:")):
            score -= 0.10
            factors.append("reply_forward_penalty")

        # Normalize score
        score = max(0.0, min(1.0, score))

        # Determine classification
        if score >= 0.7:
            classification = Classification.DATA_HIGH_CONFIDENCE
        elif score >= 0.5:
            classification = Classification.DATA_MEDIUM_CONFIDENCE
        elif score >= 0.3:
            classification = Classification.UNCERTAIN
        else:
            classification = Classification.NOT_DATA

        return ClassificationResult(
            classification=classification,
            score=score,
            factors=factors,
        )

    def _analyze_subject(self, subject: str) -> Tuple[float, List[str]]:
        """Analyze subject line for data indicators."""
        score = 0.0
        factors = []

        # High confidence patterns
        for pattern in self._high_subj:
            if pattern.search(subject):
                score += 0.30
                factors.append(f"subject_high_match:{pattern.pattern[:30]}")
                break  # Only count once

        # Medium confidence patterns
        for pattern in self._med_subj:
            if pattern.search(subject):
                score += 0.15
                factors.append(f"subject_med_match:{pattern.pattern[:30]}")
                break

        return score, factors

    def _analyze_body(self, body: str) -> Tuple[float, List[str]]:
        """Analyze body text for data delivery indicators."""
        score = 0.0
        factors = []
        body_lower = body.lower()

        # Check delivery phrases
        for phrase in self.DELIVERY_PHRASES:
            if phrase in body_lower:
                score += 0.20
                factors.append(f"body_delivery_phrase:{phrase[:20]}")
                break

        # Check metric keywords
        metric_count = 0
        for pattern in self._metric:
            if pattern.search(body):
                metric_count += 1
        if metric_count > 0:
            score += min(0.15, metric_count * 0.05)
            factors.append(f"body_metrics_mentioned:{metric_count}")

        return score, factors

    def _analyze_attachments(self, attachments: List[EmailAttachment]) -> Tuple[float, List[str]]:
        """Analyze attachments for data files."""
        score = 0.0
        factors = []

        for att in attachments:
            ext = Path(att.filename).suffix.lower()
            if ext in self.DATA_EXTENSIONS:
                score += self.DATA_EXTENSIONS[ext]
                factors.append(f"attachment_data_file:{att.filename}")
                break  # Only count once

        return score, factors

    def _analyze_links(self, body: str) -> Tuple[float, List[str]]:
        """Analyze body for data source links."""
        score = 0.0
        factors = []

        for link_type, pattern in self._links.items():
            if pattern.search(body):
                score += 0.20
                factors.append(f"link_detected:{link_type}")
                break

        return score, factors


class PatternMatcher:
    """
    Matches emails against known partner patterns.
    """

    def __init__(self, patterns: List[PartnerPattern]):
        """
        Initialize with partner patterns.

        Args:
            patterns: List of PartnerPattern configurations
        """
        self.patterns = patterns
        # Compile sender patterns
        self._compiled_patterns = []
        for p in patterns:
            compiled_senders = []
            for sender_pattern in p.sender_patterns:
                # Convert glob-style to regex
                regex = sender_pattern.replace("*", ".*").replace("?", ".")
                compiled_senders.append(re.compile(regex, re.I))

            compiled_subjects = [re.compile(sp, re.I) for sp in p.subject_patterns]
            self._compiled_patterns.append((p, compiled_senders, compiled_subjects))

    def match(
        self,
        email: EmailMessage,
        extracted_assets: List[ExtractedAsset],
    ) -> Tuple[Optional[str], float, List[str]]:
        """
        Match email against known patterns.

        Args:
            email: Email to match
            extracted_assets: Assets extracted from the email

        Returns:
            Tuple of (partner_name, confidence, matched_factors)
            partner_name is None if no match
        """
        best_match = None
        best_score = 0.0
        best_factors = []

        for pattern, sender_regexes, subject_regexes in self._compiled_patterns:
            score = 0.0
            factors = []

            # Check sender
            for regex in sender_regexes:
                if regex.search(email.from_address):
                    score += 0.30
                    factors.append("sender_match")
                    break

            # Check subject
            for regex in subject_regexes:
                if regex.search(email.subject):
                    score += 0.20
                    factors.append("subject_pattern_match")
                    break

            # Check format
            if pattern.expected_format and extracted_assets:
                for asset in extracted_assets:
                    if asset.filename:
                        ext = Path(asset.filename).suffix.lower().lstrip('.')
                        if ext == pattern.expected_format.lower().lstrip('.'):
                            score += 0.20
                            factors.append("format_match")
                            break

            # Check columns (if we have column info)
            if pattern.expected_columns:
                for asset in extracted_assets:
                    if asset.columns:
                        matched_cols = sum(
                            1 for col in pattern.expected_columns
                            if col.lower() in [c.lower() for c in asset.columns]
                        )
                        col_score = 0.30 * (matched_cols / len(pattern.expected_columns))
                        score += col_score
                        factors.append(f"column_match_{matched_cols}_of_{len(pattern.expected_columns)}")
                        break

            if score > best_score:
                best_score = score
                best_match = pattern.partner_name
                best_factors = factors

        return best_match, best_score, best_factors


class EmailScanner:
    """
    Main scanner that orchestrates email scanning, classification, and routing.
    """

    def __init__(
        self,
        gmail_client: GmailClient,
        patterns: Optional[List[PartnerPattern]] = None,
        extraction_dir: str = "/tmp/email_scanner_extracts",
        blocklist_senders: Optional[List[str]] = None,
        blocklist_domains: Optional[List[str]] = None,
    ):
        """
        Initialize the email scanner.

        Args:
            gmail_client: Authenticated GmailClient
            patterns: List of known partner patterns
            extraction_dir: Directory to save extracted attachments
            blocklist_senders: Email addresses to ignore
            blocklist_domains: Domains to ignore
        """
        self.gmail = gmail_client
        self.classifier = EmailClassifier()
        self.matcher = PatternMatcher(patterns or [])
        self.patterns = patterns or []
        self.extraction_dir = Path(extraction_dir)
        self.blocklist_senders = [s.lower() for s in (blocklist_senders or [])]
        self.blocklist_domains = [d.lower() for d in (blocklist_domains or [])]

        # Create extraction directory
        self.extraction_dir.mkdir(parents=True, exist_ok=True)

    def is_blocklisted(self, email: EmailMessage) -> bool:
        """Check if email sender is blocklisted."""
        from_addr = email.from_address.lower()

        # Check sender blocklist
        for blocked in self.blocklist_senders:
            if blocked.startswith("*@"):
                # Domain wildcard
                if from_addr.endswith(blocked[1:]):
                    return True
            elif blocked in from_addr:
                return True

        # Check domain blocklist
        for domain in self.blocklist_domains:
            if f"@{domain}" in from_addr:
                return True

        return False

    def extract_assets(
        self,
        email: EmailMessage,
        run_id: str,
    ) -> List[ExtractedAsset]:
        """
        Extract all data assets from an email.

        Args:
            email: Email to extract from
            run_id: Run ID for organizing extracts

        Returns:
            List of extracted assets
        """
        assets = []
        extract_path = self.extraction_dir / run_id

        # Extract attachments
        for att in email.attachments:
            ext = Path(att.filename).suffix.lower()

            # Only extract data files
            if ext in EmailClassifier.DATA_EXTENSIONS:
                try:
                    file_path = self.gmail.save_attachment(
                        email.message_id,
                        att,
                        str(extract_path),
                    )

                    # Try to get column info for tabular files
                    columns = self._extract_columns(file_path) if file_path.exists() else None

                    assets.append(ExtractedAsset(
                        asset_type="attachment",
                        source_email_id=email.message_id,
                        filename=att.filename,
                        file_path=file_path,
                        mime_type=att.mime_type,
                        size_bytes=att.size,
                        columns=columns,
                    ))
                except Exception as e:
                    logger.error(f"Failed to extract attachment {att.filename}: {e}")
                    assets.append(ExtractedAsset(
                        asset_type="attachment",
                        source_email_id=email.message_id,
                        filename=att.filename,
                        error=str(e),
                    ))

        # Extract links
        body = email.body_text or email.body_html or ""
        for link_type, pattern in EmailClassifier.LINK_PATTERNS.items():
            match = re.search(pattern, body, re.I)
            if match:
                assets.append(ExtractedAsset(
                    asset_type="link",
                    source_email_id=email.message_id,
                    url=match.group(0),
                    link_type=link_type,
                ))

        return assets

    def _extract_columns(self, file_path: Path) -> Optional[List[str]]:
        """Try to extract column headers from a data file."""
        try:
            ext = file_path.suffix.lower()

            if ext == '.csv':
                import csv
                with open(file_path, 'r', newline='', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    return next(reader, None)

            elif ext in ('.xlsx', '.xls'):
                # Try openpyxl for xlsx
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True)
                    ws = wb.active
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    return [str(c) if c else "" for c in first_row]
                except ImportError:
                    pass

            elif ext == '.json':
                import json
                with open(file_path, 'r') as f:
                    data = json.load(f)
                    if isinstance(data, list) and data:
                        return list(data[0].keys()) if isinstance(data[0], dict) else None
                    elif isinstance(data, dict):
                        return list(data.keys())

        except Exception as e:
            logger.debug(f"Could not extract columns from {file_path}: {e}")

        return None

    def scan_email(self, email: EmailMessage, run_id: str) -> ScanResult:
        """
        Scan a single email.

        Args:
            email: Email to scan
            run_id: Run ID for this scan

        Returns:
            ScanResult with classification and extracted assets
        """
        # Check blocklist first
        if self.is_blocklisted(email):
            return ScanResult(
                email=email,
                classification=ClassificationResult(
                    classification=Classification.NOT_DATA,
                    score=0.0,
                    factors=["blocklisted_sender"],
                ),
                extracted_assets=[],
                route_action="skip",
            )

        # Classify email
        classification = self.classifier.classify(email)

        # If not data, skip extraction
        if classification.classification == Classification.NOT_DATA:
            return ScanResult(
                email=email,
                classification=classification,
                extracted_assets=[],
                route_action="skip",
            )

        # Extract assets
        assets = self.extract_assets(email, run_id)

        # Match against patterns
        partner, pattern_confidence, match_factors = self.matcher.match(email, assets)

        # Determine routing
        route_action = "review"  # Default
        if partner:
            # Find the pattern config
            for p in self.patterns:
                if p.partner_name == partner:
                    if p.auto_process and pattern_confidence >= p.confidence_threshold:
                        route_action = "auto_process"
                    break
        elif classification.classification == Classification.NOT_DATA:
            route_action = "skip"

        # Update classification factors with pattern info
        classification.matched_patterns = match_factors

        return ScanResult(
            email=email,
            classification=classification,
            extracted_assets=assets,
            matched_partner=partner,
            pattern_match_confidence=pattern_confidence,
            route_action=route_action,
        )

    def scan(
        self,
        since: Optional[datetime] = None,
        query: Optional[str] = None,
        max_results: int = 100,
        labels: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Run a full scan.

        Args:
            since: Only scan emails after this time
            query: Gmail search query
            max_results: Maximum emails to process
            labels: Filter by labels

        Returns:
            Scan results dictionary matching contract output format
        """
        run_id = f"es-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        started_at = datetime.now()

        results = {
            "run_id": run_id,
            "started_at": started_at.isoformat(),
            "emails_scanned": 0,
            "emails_classified_as_data": 0,
            "auto_processed": [],
            "review_queue": [],
            "errors": [],
        }

        for email in self.gmail.fetch_emails(
            since=since,
            query=query,
            max_results=max_results,
            labels=labels,
        ):
            results["emails_scanned"] += 1

            try:
                scan_result = self.scan_email(email, run_id)

                if scan_result.classification.classification != Classification.NOT_DATA:
                    results["emails_classified_as_data"] += 1

                if scan_result.route_action == "auto_process":
                    results["auto_processed"].append(
                        self._format_auto_process_payload(scan_result, run_id)
                    )
                elif scan_result.route_action == "review":
                    results["review_queue"].append(
                        self._format_review_item(scan_result, run_id)
                    )

            except Exception as e:
                logger.error(f"Error scanning email {email.message_id}: {e}")
                results["errors"].append({
                    "email_id": email.message_id,
                    "error": str(e),
                })

        results["completed_at"] = datetime.now().isoformat()
        return results

    def _format_auto_process_payload(self, result: ScanResult, run_id: str) -> Dict[str, Any]:
        """Format scan result as auto-process payload for Data Harmonization."""
        assets = []
        for asset in result.extracted_assets:
            if asset.file_path:
                assets.append({
                    "source_system": "email",
                    "source_location": f"email://{result.email.message_id}/{asset.filename}",
                    "partner_name": result.matched_partner,
                    "received_at": result.email.date.isoformat(),
                    "payload": {
                        "type": "file_reference",
                        "path": str(asset.file_path),
                    },
                    "email_metadata": {
                        "message_id": result.email.message_id,
                        "from": result.email.from_address,
                        "subject": result.email.subject,
                        "received_date": result.email.date.isoformat(),
                    },
                })

        return {
            "payload_id": f"{run_id}-{result.email.message_id[:8]}",
            "partner_name": result.matched_partner,
            "pattern_match_confidence": result.pattern_match_confidence,
            "raw_sources": assets,
        }

    def _format_review_item(self, result: ScanResult, run_id: str) -> Dict[str, Any]:
        """Format scan result as review queue item."""
        return {
            "review_id": f"rev-{run_id}-{result.email.message_id[:8]}",
            "email_summary": {
                "message_id": result.email.message_id,
                "from": result.email.from_address,
                "subject": result.email.subject,
                "date": result.email.date.isoformat(),
                "snippet": result.email.snippet,
            },
            "classification_score": result.classification.score,
            "classification_factors": result.classification.factors,
            "extracted_assets": [
                {
                    "type": a.asset_type,
                    "filename": a.filename,
                    "url": a.url,
                    "columns": a.columns,
                    "error": a.error,
                }
                for a in result.extracted_assets
            ],
            "suggested_partner": result.matched_partner,
            "pattern_match_confidence": result.pattern_match_confidence,
        }
